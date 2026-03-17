import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATA_FILE = Path(__file__).resolve().parent.parent / "data" / "project_cases.jsonl"
EXCEL_FILE = Path(__file__).resolve().parent.parent / "data" / "\u8d85\u4e1a\u8d44\u8d282026.xlsx"
DEFAULT_SHEET_NAME: str | None = None

COL_PROJECT_NAME = "\u9879\u76ee\u540d\u79f0"
COL_COMPANY_NAME = "\u5efa\u8bbe\u5355\u4f4d"
COL_PROJECT_DATE = "\u5f00\u7ae3\u5de5\u65e5\u671f"
COL_PROJECT_AMOUNT = "\u5408\u540c\u4ef7\u683c(\u4e07\u5143)"
COL_LOCATION = "\u5efa\u8bbe\u5730\u70b9"
REQUIRED_HEADERS = {
    COL_PROJECT_NAME,
    COL_COMPANY_NAME,
    COL_PROJECT_DATE,
    COL_PROJECT_AMOUNT,
    COL_LOCATION,
}


def _parse_json_objects(text: str, file_path: str, line_number: int) -> list[dict]:
    """Parse one or more JSON objects from a line."""
    decoder = json.JSONDecoder()
    index = 0
    items: list[dict] = []

    while index < len(text):
        while index < len(text) and text[index].isspace():
            index += 1
        if index >= len(text):
            break
        try:
            item, next_index = decoder.raw_decode(text, index)
        except json.JSONDecodeError as exc:
            raise ValueError(
                f"Invalid JSON on line {line_number} in {file_path}: {exc}"
            ) from exc
        if not isinstance(item, dict):
            raise ValueError(
                f"Invalid record on line {line_number} in {file_path}: expected object"
            )
        items.append(item)
        index = next_index

    return items



def _clean_header(value: Any) -> str:
    """Normalize Excel headers for matching."""
    text = str(value or "").replace("\n", "").replace(" ", "").strip()
    return text



def load_existing_cases(file_path: str) -> list[dict]:
    """Load existing cases from the JSONL file."""
    path = Path(file_path)
    if not path.exists():
        return []

    cases: list[dict] = []
    with path.open("r", encoding="utf-8") as file:
        for line_number, line in enumerate(file, start=1):
            content = line.strip()
            if not content:
                continue
            cases.extend(_parse_json_objects(content, file_path, line_number))
    return cases



def read_excel_rows(excel_path: str, sheet_name: str | None = None) -> list[dict]:
    """Read row dictionaries from an Excel worksheet."""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_index = None
    headers: list[str] = []
    for index, raw_row in enumerate(rows):
        cleaned = [_clean_header(value) for value in raw_row]
        if REQUIRED_HEADERS.issubset({item for item in cleaned if item}):
            header_row_index = index
            headers = cleaned
            break

    if header_row_index is None:
        raise ValueError("Required Excel headers were not found.")

    result: list[dict] = []
    for values in rows[header_row_index + 1 :]:
        row = {
            headers[index]: values[index] if index < len(values) else None
            for index in range(len(headers))
            if headers[index]
        }
        if any(value not in (None, "") for value in row.values()):
            result.append(row)
    return result



def extract_business_type_and_keywords(project_name: str) -> tuple[str, list[str]]:
    """Extract a simple business type and keyword list from project name."""
    keyword_map = [
        ("\u9ad8\u4f4e\u538b\u914d\u7535", "\u9ad8\u4f4e\u538b\u914d\u7535\u5de5\u7a0b", "\u9ad8\u4f4e\u538b\u914d\u7535"),
        ("\u53d8\u914d\u7535", "\u53d8\u914d\u7535\u5de5\u7a0b", "\u53d8\u914d\u7535"),
        ("\u914d\u7535\u5b89\u88c5", "\u914d\u7535\u5b89\u88c5\u5de5\u7a0b", "\u914d\u7535\u5b89\u88c5"),
        ("\u914d\u7535\u5de5\u7a0b", "\u914d\u7535\u5de5\u7a0b", "\u914d\u7535\u5de5\u7a0b"),
        ("\u914d\u7535", "\u914d\u7535\u5de5\u7a0b", "\u914d\u7535"),
        ("\u7535\u529b", "\u7535\u529b\u5de5\u7a0b", "\u7535\u529b"),
        ("\u5b89\u88c5", "\u5b89\u88c5\u5de5\u7a0b", "\u5b89\u88c5"),
    ]

    matched_keywords: list[str] = []
    business_type = ""
    for needle, inferred_type, keyword in keyword_map:
        if needle in project_name:
            if not business_type:
                business_type = inferred_type
            if keyword not in matched_keywords:
                matched_keywords.append(keyword)

    if not business_type:
        business_type = "\u9879\u76ee\u5bfc\u5165"
    return business_type, matched_keywords



def _extract_city(location_text: str) -> str:
    """Extract a rough city name from location text."""
    text = str(location_text or "").strip()
    if not text:
        return ""

    city_marker = "\u5e02"
    province_marker = "\u7701"
    if city_marker in text:
        return text[: text.find(city_marker) + 1]
    if province_marker in text and len(text) > text.find(province_marker) + 1:
        return text[text.find(province_marker) + 1 :]
    if text.startswith("\u6df1\u5733"):
        return "\u6df1\u5733"
    if text.startswith("\u4e1c\u839e"):
        return "\u4e1c\u839e"
    if text.startswith("\u5e7f\u5dde"):
        return "\u5e7f\u5dde"
    if text.startswith("\u60e0\u5dde"):
        return "\u60e0\u5dde"
    return text



def _format_cell_value(value: Any) -> str:
    """Format cell values as strings."""
    if value is None:
        return ""
    return str(value).replace("\n", "").strip()



def _convert_amount_to_yuan(value: Any) -> int:
    """Convert contract price in ten-thousand yuan to yuan."""
    if value in (None, ""):
        return 0
    try:
        amount_wan = float(value)
    except (TypeError, ValueError):
        return 0
    return int(amount_wan * 10000)



def normalize_project_row(row: dict, auto_index: int) -> dict:
    """Normalize one Excel row into the standard case schema."""
    project_name = _format_cell_value(row.get(COL_PROJECT_NAME))
    company_name = _format_cell_value(row.get(COL_COMPANY_NAME))
    business_type, keywords = extract_business_type_and_keywords(project_name)
    location_text = _format_cell_value(row.get(COL_LOCATION))
    project_date_raw = _format_cell_value(row.get(COL_PROJECT_DATE))

    return {
        "project_id": f"AUTO_{auto_index:04d}",
        "project_name": project_name,
        "company_name": company_name,
        "industry": "",
        "business_type": business_type,
        "location_city": _extract_city(location_text),
        "location_district": "",
        "project_amount": _convert_amount_to_yuan(row.get(COL_PROJECT_AMOUNT)),
        "customer_problem": "",
        "solution_summary": "",
        "project_stage": "",
        "owner_role": "",
        "duration_estimate": "",
        "risk_notes": [],
        "keywords": keywords,
        "custom_fields": {
            "project_date_raw": project_date_raw,
            "source_construction_location": location_text,
        },
    }



def case_exists(existing_cases: list[dict], project_name: str, company_name: str) -> bool:
    """Check whether the project already exists by name and company."""
    target_name = str(project_name).strip()
    target_company = str(company_name).strip()
    for case in existing_cases:
        if (
            str(case.get("project_name", "")).strip() == target_name
            and str(case.get("company_name", "")).strip() == target_company
        ):
            return True
    return False



def append_cases_to_jsonl(file_path: str, cases: list[dict]) -> None:
    """Append normalized cases to the JSONL file."""
    if not cases:
        return
    path = Path(file_path)
    prefix = ""
    if path.exists() and path.stat().st_size > 0:
        prefix = "\n"
    with path.open("a", encoding="utf-8") as file:
        for index, case in enumerate(cases):
            if index == 0:
                file.write(prefix)
            else:
                file.write("\n")
            file.write(json.dumps(case, ensure_ascii=False))



def main() -> None:
    """Import Excel ledger rows into project_cases.jsonl."""
    excel_path = str(EXCEL_FILE)
    sheet_name = DEFAULT_SHEET_NAME

    try:
        existing_cases = load_existing_cases(str(DATA_FILE))
        excel_rows = read_excel_rows(excel_path, sheet_name)

        total_rows = len(excel_rows)
        skipped_count = 0
        cases_to_write: list[dict] = []
        seen_pairs = {
            (
                str(case.get("project_name", "")).strip(),
                str(case.get("company_name", "")).strip(),
            )
            for case in existing_cases
        }

        next_auto_index = 1
        for case in existing_cases:
            project_id = str(case.get("project_id", "")).strip()
            if project_id.startswith("AUTO_"):
                suffix = project_id.removeprefix("AUTO_")
                if suffix.isdigit():
                    next_auto_index = max(next_auto_index, int(suffix) + 1)

        for row in excel_rows:
            project_name = _format_cell_value(row.get(COL_PROJECT_NAME))
            company_name = _format_cell_value(row.get(COL_COMPANY_NAME))
            if not project_name:
                skipped_count += 1
                continue
            if case_exists(existing_cases, project_name, company_name) or (
                project_name,
                company_name,
            ) in seen_pairs:
                skipped_count += 1
                continue

            normalized_case = normalize_project_row(row, next_auto_index)
            next_auto_index += 1
            cases_to_write.append(normalized_case)
            seen_pairs.add((project_name, company_name))

        append_cases_to_jsonl(str(DATA_FILE), cases_to_write)

        print("\u672c\u6b21\u5bfc\u5165\u603b\u884c\u6570: " + str(total_rows))
        print("\u6210\u529f\u5199\u5165\u6761\u6570: " + str(len(cases_to_write)))
        print("\u8df3\u8fc7\u6761\u6570: " + str(skipped_count))
    except (FileNotFoundError, ValueError, OSError) as exc:
        print("\u6267\u884c\u5931\u8d25: " + str(exc))


if __name__ == "__main__":
    main()
